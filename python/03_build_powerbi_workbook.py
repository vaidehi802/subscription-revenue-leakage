"""
Assemble all analytical datasets into one formatted Power BI-ready workbook.
Each view becomes a sheet; a cover sheet documents the model.
"""
import os, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

BASE = os.path.join(os.path.dirname(__file__), "..")
PBI  = os.path.join(BASE, "powerbi", "datasets")
OUT  = os.path.join(BASE, "powerbi", "Subscription_Revenue_Leakage_Data.xlsx")

NAVY   = "1F3864"
BLUE   = "2E5496"
LIGHT  = "D9E1F2"
WHITE  = "FFFFFF"
GREY   = "F2F2F2"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_sheet(ws, df, title):
    # Title row
    ws.insert_rows(1, 2)
    ncol = max(len(df.columns), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    tcell = ws.cell(row=1, column=1, value=title)
    tcell.font = Font(name="Arial", size=13, bold=True, color=WHITE)
    tcell.fill = PatternFill("solid", fgColor=NAVY)
    tcell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    header_row = 3
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=header_row, column=j, value=str(col))
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[header_row].height = 30

    for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=header_row+1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = Font(name="Arial", size=10)
            c.border = border
            if i % 2 == 0:
                c.fill = PatternFill("solid", fgColor=GREY)
            colname = str(df.columns[j-1]).lower()
            if any(k in colname for k in ["revenue","mrr","arr","loss","amount","clv","discount","value","lost","risk"]):
                if isinstance(val,(int,float)):
                    c.number_format = '#,##0'
                    c.alignment = Alignment(horizontal="right")
            elif "pct" in colname or "rate" in colname:
                if isinstance(val,(int,float)):
                    c.number_format = '0.0'
                    c.alignment = Alignment(horizontal="right")

    # column widths
    for j, col in enumerate(df.columns, start=1):
        maxlen = max([len(str(col))] + [len(str(v)) for v in df.iloc[:,j-1].astype(str)])
        ws.column_dimensions[get_column_letter(j)].width = min(max(maxlen+3, 12), 40)
    ws.freeze_panes = ws.cell(row=header_row+1, column=1)
    ws.sheet_view.showGridLines = False

# Map view files -> sheet name + section title
sheets = [
    ("v_kpi_summary.csv",            "KPI_Summary",        "Dashboard KPIs — Executive Summary"),
    ("v_revenue_overview.csv",       "1_Revenue_Overview", "1. Revenue Overview"),
    ("v_failed_payments.csv",        "2_Failed_Monthly",   "2a. Failed Payments — Monthly Trend"),
    ("v_failed_payment_by_plan.csv", "2_Failed_By_Plan",   "2b. Failed Payments by Plan"),
    ("v_failed_payment_reasons.csv", "2_Failed_Reasons",   "2c. Failed Payment Reasons"),
    ("v_churn_monthly.csv",          "3_Churn_Monthly",    "3a. Churn — Monthly Trend"),
    ("v_churn_reasons.csv",          "3_Churn_Reasons",    "3b. Churn Reasons"),
    ("v_expired_subscriptions.csv",  "4_Expired",          "4. Expired Subscription Analysis"),
    ("v_downgrade_analysis.csv",     "5_Downgrades",       "5. Downgrade Analysis"),
    ("v_discount_leakage.csv",       "6_Discount_Leakage", "6. Discount Leakage Analysis"),
    ("v_retention_metrics.csv",      "7_Retention",        "7. Customer Retention Metrics"),
    ("v_leakage_root_cause.csv",     "8_Leakage_RootCause","8. Revenue Leakage Root Cause"),
    ("v_cohort_retention.csv",       "9_Cohort",           "9. Cohort Retention Analysis"),
]

wb = Workbook()
# Cover sheet
cover = wb.active
cover.title = "Cover"
cover.sheet_view.showGridLines = False
cover.merge_cells("A1:F1")
c = cover["A1"]; c.value = "Subscription Revenue Leakage — Analytics Dataset"
c.font = Font(name="Arial", size=18, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")
cover.row_dimensions[1].height = 40
cover["A3"] = "Pipeline: Python (data) -> SQL/SQLite (analysis views) -> Power BI (this workbook)"
cover["A4"] = "Each sheet maps to a dashboard section. Import this workbook into Power BI Desktop (Get Data > Excel)."
cover["A3"].font = cover["A4"].font = Font(name="Arial", size=10)
cover["A6"] = "Sheet"; cover["B6"] = "Dashboard Section"
for cc in ["A6","B6"]:
    cover[cc].font = Font(name="Arial", bold=True, color=WHITE)
    cover[cc].fill = PatternFill("solid", fgColor=BLUE)
r = 7
for _, name, title in sheets:
    cover[f"A{r}"] = name; cover[f"B{r}"] = title
    cover[f"A{r}"].font = cover[f"B{r}"].font = Font(name="Arial", size=10)
    r += 1
cover.column_dimensions["A"].width = 24
cover.column_dimensions["B"].width = 50

for fname, sheet_name, title in sheets:
    df = pd.read_csv(os.path.join(PBI, fname))
    ws = wb.create_sheet(sheet_name)
    # write df starting at A1, style adds title rows
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    # remove the appended rows; restyle cleanly
    ws.delete_rows(1, ws.max_row)
    style_sheet(ws, df, title)

wb.save(OUT)
print(f"Saved workbook: {OUT}")
print(f"Sheets: {wb.sheetnames}")
